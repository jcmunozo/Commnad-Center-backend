"""Excel → relational import orchestrator.

Reads the PMO Command Center workbook, normalizes each sheet's headers to the
target Resource field names, and runs django-import-export Resources in FK-safe
order. Supports dry-run (validate only) and confirm (persist).

The whole pipeline runs inside ONE transaction with per-resource transactions
disabled, so rows persisted by an earlier sheet are visible to later sheets'
FK lookups (Projects → Team, APIs → Projects, …). A dry-run executes the same
work and rolls the transaction back at the end, so its report matches exactly
what a confirm would do.

The COLUMN_MAP layer bridges the Excel's human headers ("Project ID") to the
Resource fields ("legacy_code"). Sheets not listed are skipped. Team runs in
two passes because Employee.manager is a self-reference resolved by name: rows
are created first without it, then a second pass fills it in.
"""
from __future__ import annotations

import tablib
from django.db import DatabaseError, transaction
from openpyxl import load_workbook

from apps.clients.models import Client
from apps.projects.resources import (
    ApiComponentResource,
    EndpointResource,
    MilestoneResource,
    ProjectResource,
    TaskResource,
)
from apps.resources.resources import EmployeeResource, TaskAssignmentResource
from apps.tracking.resources import ActionResource, IssueResource, RiskResource

# Ordered: catalogs are seeded separately (seed_catalogs); FKs resolve top-down.
PIPELINE = [
    # (sheet_name, report_label, Resource, {excel_header: resource_field})
    ("Team", "Team", EmployeeResource, {
        "Employee ID": "legacy_code", "Name": "name", "Role": "role",
        "Availability %": "availability_pct", "Weekly Hrs": "weekly_hours",
        "Level": "level", "Status": "status", "Location": "location", "Timezone": "timezone",
    }),
    ("Team", "Team (managers)", EmployeeResource, {
        "Employee ID": "legacy_code", "Manager": "manager",
    }),
    ("Projects", "Projects", ProjectResource, {
        "Project ID": "legacy_code", "Project Name": "name", "Client": "client",
        "Start Date": "start_date", "Planned End": "planned_end", "Actual End": "actual_end",
        "Status": "status", "Priority": "priority", "Health": "health",
        "Progress %": "progress_pct", "Planned Hrs": "planned_hours",
        "Consumed Hrs": "consumed_hours", "Comments": "comments",
    }),
    ("APIs", "APIs", ApiComponentResource, {
        "API ID": "legacy_code", "Project ID": "owner_project", "API Name": "name",
        "Description": "description", "Version": "version",
        "Status": "status", "Owner": "owner_employee", "Comments": "comments",
    }),
    ("Endpoints", "Endpoints", EndpointResource, {
        "Endpoint ID": "legacy_code", "API ID": "api", "HTTP Method": "http_method",
        "Path": "path", "Description": "description",
        "Status": "status", "Owner": "owner_employee", "Comments": "comments",
    }),
    ("Tasks", "Tasks", TaskResource, {
        "Task ID": "legacy_code", "Project ID": "project", "API ID": "api",
        "Endpoint ID": "endpoint", "Task Type": "task_type", "Task Name": "name",
        "Planned Start": "planned_start", "Planned End": "planned_end",
        "Status": "status", "Priority": "priority", "Estimated Hrs": "estimated_hours",
        "Actual Hrs": "actual_hours", "Progress %": "progress_pct", "Notes": "notes",
    }),
    ("Milestones", "Milestones", MilestoneResource, {
        "Milestone ID": "legacy_code", "Project ID": "project", "API ID": "api",
        "Milestone Name": "name", "Owner": "owner_employee",
        "Target Date": "target_date", "Actual Date": "actual_date", "Comments": "comments",
    }),
    ("Assignments", "Assignments", TaskAssignmentResource, {
        "Assignment ID": "legacy_code", "Task ID": "task", "Employee ID": "employee",
        "Assigned Date": "assigned_date", "Delivery Date": "delivery_date",
    }),
    ("Issues", "Issues", IssueResource, {
        "Issue ID": "legacy_code", "Project ID": "project", "API ID": "api",
        "Date Reported": "date_reported", "Reported By": "reported_by_name",
        "Description": "description", "Impact": "impact", "Urgency": "urgency",
        "Assignee": "assignee", "Status": "status",
        "Resolution Date": "resolution_date", "Lessons Learned": "lessons_learned",
    }),
    ("Risks", "Risks", RiskResource, {
        "Risk ID": "legacy_code", "Project ID": "project", "Description": "description",
        "Probability": "probability", "Impact": "impact",
        "Mitigation Plan": "mitigation_plan", "Owner": "owner_employee", "Status": "status",
        "Identified Date": "identified_date", "Last Review": "last_review",
    }),
    ("Actions", "Actions", ActionResource, {
        "Action ID": "legacy_code", "Created": "created_date", "Project ID": "project",
        "API ID": "api", "Origin": "origin", "Description": "description",
        "Assignee": "assignee", "Due Date": "due_date", "Priority": "priority",
        "Status": "status", "Impact": "impact", "Notes": "notes",
    }),
]


def _sheet_to_dataset(ws, column_map: dict[str, str]) -> tablib.Dataset:
    """Convert a worksheet to a Dataset, keeping only mapped columns, skipping
    rows whose ID (first mapped column) is blank (filters formula-only rows)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return tablib.Dataset()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    keep = [(i, column_map[h]) for i, h in enumerate(headers) if h in column_map]
    ds = tablib.Dataset(headers=[field for _, field in keep])
    id_field = keep[0][1] if keep else None
    for raw in rows[1:]:
        values = [raw[i] if i < len(raw) else None for i, _ in keep]
        if id_field and (values[0] is None or str(values[0]).strip() == ""):
            continue  # skip pre-seeded formula rows with no real ID
        ds.append(values)
    return ds


def _ensure_clients(wb) -> int:
    """The workbook has no Clients sheet: create Client rows from the distinct
    names in the Projects sheet so ProjectResource's by-name lookup resolves."""
    if "Projects" not in wb.sheetnames:
        return 0
    rows = wb["Projects"].iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows, [])]
    if "Client" not in headers:
        return 0
    idx = headers.index("Client")
    created = 0
    seen = set()
    for raw in rows:
        name = raw[idx] if idx < len(raw) else None
        name = str(name).strip() if name is not None else ""
        if not name or name in seen:
            continue
        seen.add(name)
        _, was_created = Client.objects.get_or_create(name=name)
        created += int(was_created)
    return created


def run_import(file_obj, *, dry_run: bool) -> dict:
    """Execute the pipeline. Returns a per-entity report with row-level errors."""
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    report = {"dry_run": dry_run, "entities": [], "has_errors": False}

    with transaction.atomic():
        report["clients_created"] = _ensure_clients(wb)

        for sheet_name, label, resource_cls, column_map in PIPELINE:
            if sheet_name not in wb.sheetnames:
                continue
            dataset = _sheet_to_dataset(wb[sheet_name], column_map)
            resource = resource_cls()
            # dry_run=False + use_transactions=False: rows persist inside OUR
            # transaction so later sheets can resolve them; the final rollback
            # (dry-run) or the confirm view's outer atomic decides their fate.
            # The per-sheet savepoint exists to recover from a DB-level error
            # (which aborts the postgres transaction until rolled back).
            sid = transaction.savepoint()
            result = resource.import_data(
                dataset, dry_run=False, use_transactions=False, raise_errors=False)

            row_errors = []
            db_error = False
            for line, errors in result.row_errors():
                row_errors.append({"row": line, "errors": [str(e.error) for e in errors]})
                db_error = db_error or any(isinstance(e.error, DatabaseError) for e in errors)
            if db_error:
                transaction.savepoint_rollback(sid)
            else:
                transaction.savepoint_commit(sid)
            for invalid in getattr(result, "invalid_rows", []):
                row_errors.append({
                    "row": invalid.number,
                    "errors": [f"{k}: {v}" for k, v in invalid.error_dict.items()],
                })

            entity_report = {
                "sheet": label,
                "model": resource_cls.Meta.model.__name__,
                "total_rows": len(dataset),
                "new": result.totals.get("new", 0),
                "updated": result.totals.get("update", 0),
                "skipped": result.totals.get("skip", 0),
                "invalid": result.totals.get("invalid", 0),
                "row_errors": row_errors,
            }
            if row_errors or entity_report["invalid"]:
                report["has_errors"] = True
            report["entities"].append(entity_report)

        if dry_run or report["has_errors"]:
            transaction.set_rollback(True)

    return report
